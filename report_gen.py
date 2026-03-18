"""
report_gen.py — Excel and PDF report generation for PMCC trade history
"""
import io
import os
from datetime import datetime
from typing import Optional

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from fpdf import FPDF
    FPDF_OK = True
except ImportError:
    FPDF_OK = False


# ─── Excel ────────────────────────────────────────────────────────────────────

def generate_excel(trades_df: pd.DataFrame,
                   pnl_df: Optional[pd.DataFrame] = None) -> bytes:
    """
    Produce a styled Excel workbook as bytes.
    Returns raw bytes suitable for st.download_button.
    """
    if not OPENPYXL_OK:
        return _csv_fallback(trades_df)

    wb = openpyxl.Workbook()

    # ── Trades Sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Trades"

    header_fill = PatternFill("solid", fgColor="0A0E1A")
    header_font = Font(color="00D4FF", bold=True, name="Calibri", size=11)
    alt_fill    = PatternFill("solid", fgColor="111827")
    border_side = Side(style="thin", color="1E293B")
    thin_border = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )

    cols = list(trades_df.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, row in enumerate(dataframe_to_rows(trades_df, index=False, header=False)):
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="0D1526")
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # ── PnL Summary Sheet ────────────────────────────────────────────────────
    if pnl_df is not None and not pnl_df.empty:
        ws2 = wb.create_sheet("PnL Summary")
        ws2.append(list(pnl_df.columns))
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in dataframe_to_rows(pnl_df, index=False, header=False):
            ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_fallback(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


# ─── PDF ──────────────────────────────────────────────────────────────────────

def generate_pdf(trades_df: pd.DataFrame,
                 pnl_df: Optional[pd.DataFrame] = None) -> bytes:
    """
    Produce a styled PDF report as bytes.
    """
    if not FPDF_OK:
        return _csv_fallback(trades_df)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_fill_color(10, 14, 26)        # bg dark
    pdf.rect(0, 0, 297, 210, "F")

    # ── Title ─────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(0, 212, 255)        # accent cyan
    pdf.cell(0, 12, "PMCC Quant-Dashboard - Trade Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
             ln=True, align="C")
    pdf.ln(6)

    # ── PnL Summary ───────────────────────────────────────────────────────────
    if pnl_df is not None and not pnl_df.empty:
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(16, 185, 129)
        pdf.cell(0, 8, "PnL Summary", ln=True)
        pdf.ln(2)
        _pdf_table(pdf, pnl_df)
        pdf.ln(8)

    # ── Trades ────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(0, 212, 255)
    pdf.cell(0, 8, "Trade History", ln=True)
    pdf.ln(2)
    _pdf_table(pdf, trades_df.head(50))   # cap to 50 rows for readability

    buf = io.BytesIO()
    pdf_bytes = pdf.output()
    return bytes(pdf_bytes)


def _pdf_table(pdf: "FPDF", df: pd.DataFrame) -> None:
    if df.empty:
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 6, "(No data)", ln=True)
        return

    cols = list(df.columns)
    col_w = min(40, 260 // max(len(cols), 1))

    # Header
    pdf.set_fill_color(17, 24, 39)
    pdf.set_text_color(0, 212, 255)
    pdf.set_font("Helvetica", "B", 8)
    for col in cols:
        pdf.cell(col_w, 7, str(col)[:14], border=1, align="C", fill=True)
    pdf.ln()

    # Rows
    pdf.set_font("Helvetica", "", 7)
    for i, (_, row) in enumerate(df.iterrows()):
        fill = i % 2 == 0
        if fill:
            pdf.set_fill_color(13, 21, 38)
        else:
            pdf.set_fill_color(17, 24, 39)
        pdf.set_text_color(226, 232, 240)
        for col in cols:
            val = str(row[col])[:16]
            pdf.cell(col_w, 6, val, border=1, align="C", fill=True)
        pdf.ln()
